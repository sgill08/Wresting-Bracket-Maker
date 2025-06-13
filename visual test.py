import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import Workbook

# Initialize the main application
# Ayden Winter
class WrestlingBracketApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Wrestling Bracket Software")
        self.root.geometry("800x600")

        # Variables
        self.input_file = None
        self.data = None
        self.bracket_type = tk.StringVar(value="Round Robin")
        self.tournament_type = tk.StringVar(value="Individual")
        self.num_mats = tk.IntVar(value=1)

        # Layout
        self.create_widgets()

    # Seth G
    def create_widgets(self):
        # Frame for file selection
        file_frame = ttk.Frame(self.root, padding="10")
        file_frame.pack(fill=tk.X)

        self.file_label = ttk.Label(file_frame, text="No file selected")
        self.file_label.pack(side=tk.LEFT, padx=5)

        select_file_btn = ttk.Button(file_frame, text="Select File", command=self.select_file)
        select_file_btn.pack(side=tk.RIGHT, padx=5)

        # Frame for user inputs
        input_frame = ttk.LabelFrame(self.root, text="Tournament Settings", padding="10")
        input_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(input_frame, text="Bracket Type:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Combobox(input_frame, textvariable=self.bracket_type, values=["Round Robin", "Double Elimination"], state="readonly").grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(input_frame, text="Tournament Type:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Combobox(input_frame, textvariable=self.tournament_type, values=["Dual Meet", "Individual"], state="readonly").grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(input_frame, text="Number of Mats:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Spinbox(input_frame, from_=1, to=16, textvariable=self.num_mats, width=5).grid(row=2, column=1, padx=5, pady=5)

        # Frame for data preview
        # Ayden Winter
        preview_frame = ttk.LabelFrame(self.root, text="Data Preview", padding="10")
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(preview_frame, columns=("Last Name", "First Name", "Weight", "School", "Rank", "Gender"), show="headings")
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        self.tree.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Seth G
        # Frame for actions
        action_frame = ttk.Frame(self.root, padding="10")
        action_frame.pack(fill=tk.X)

        process_btn = ttk.Button(action_frame, text="Process Data", command=self.process_data)
        process_btn.pack(side=tk.LEFT, padx=5)

        edit_btn = ttk.Button(action_frame, text="Edit Selected", command=self.edit_selected)
        edit_btn.pack(side=tk.LEFT, padx=5)

        delete_btn = ttk.Button(action_frame, text="Delete Selected", command=self.delete_selected)
        delete_btn.pack(side=tk.LEFT, padx=5)

        save_btn = ttk.Button(action_frame, text="Save Changes", command=self.save_changes)
        save_btn.pack(side=tk.LEFT, padx=5)

        exit_btn = ttk.Button(action_frame, text="Exit", command=self.root.quit)
        exit_btn.pack(side=tk.RIGHT, padx=5)

        # New frame for the "Generate Brackets" button
        bottom_frame = ttk.Frame(self.root, padding="10")
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X)

        generate_btn = ttk.Button(bottom_frame, text="Generate Brackets", command=self.generate_brackets)
        generate_btn.pack(pady=5)
         
    # Ayden Winter
    def generate_brackets(self):
        if self.data is None:
            messagebox.showwarning("Warning", "No data loaded. Please select and process a file first.")
        else:
            try:
                # Sort wrestlers by gender, rank, and weight
                sorted_data = self.data.sort_values(by=['Gender', 'Rank', 'Weight'], ascending=[True, True, False])

                # Create a new Excel workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Brackets"

                # Variables for bracket generation
                mat_number = 1
                bracket_counter = 0
                weight_class = "Any"  # You can adjust this if you have weight class data

                # Start filling the Excel template
                row_idx = 2  # Start from the second row for data population

                for gender in ['boy', 'girl']:  # Process boys and girls separately
                    gender_data = sorted_data[sorted_data['Gender'].str.lower() == gender]
                
                    while len(gender_data) > 0:
                        bracket = []
                        bracket_counter += 1

                        # Take up to 4 wrestlers for the bracket
                        bracket = gender_data.head(4)
                        gender_data = gender_data.iloc[4:]

                        # Fill in the template for the current bracket
                        ws[f"B{row_idx}"] = weight_class
                        ws[f"E{row_idx}"] = mat_number

                        for i, wrestler in enumerate(bracket.iterrows()):
                            wrestler_data = wrestler[1]

                            # Convert all values to strings before populating
                            last_name = str(wrestler_data['Last Name']) if pd.notna(wrestler_data['Last Name']) else ""
                            first_name = str(wrestler_data['First Name']) if pd.notna(wrestler_data['First Name']) else ""
                            full_name = last_name + ", " + first_name  # Concatenate last and first names
                        
                            school = str(wrestler_data['School']) if pd.notna(wrestler_data['School']) else "Unknown"
                            rank = str(wrestler_data['Rank']) if pd.notna(wrestler_data['Rank']) else "N/A"
                            gender = str(wrestler_data['Gender']) if pd.notna(wrestler_data['Gender']) else "Unknown"
                        
                            # Write to Excel cells
                            ws[f"B{row_idx + 1 + i}"] = full_name
                            ws[f"C{row_idx + 1 + i}"] = school
                            ws[f"D{row_idx + 1 + i}"] = rank
                            ws[f"E{row_idx + 1 + i}"] = gender

                        mat_number += 1
                        row_idx += len(bracket) + 1  # Move to next row after each bracket

                        # Restart mat numbers after max mats are reached
                        if mat_number > self.num_mats.get():
                            mat_number = 1

                # Save the Excel file
                output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
                if output_file:
                    wb.save(output_file)

                    messagebox.showinfo("Success", "Brackets generated and saved successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate brackets: {e}")

    # Ayden Winter
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File", 
            filetypes=(
                ("Excel Files", "*.xlsx"),
                ("All Files", "*.*")
            )
        )

        if file_path:
            self.input_file = file_path
            self.file_label.config(text=f"Selected: {file_path.split('/')[-1]}")
            self.load_data()
    # Ayden Winter
    def load_data(self):
        try:
            # Load Excel sheets dynamically
            sheets = pd.ExcelFile(self.input_file).sheet_names
            all_data = []

            for sheet in sheets:
                try:
                    df = pd.read_excel(self.input_file, sheet_name=sheet, skiprows=2)
                
                    if not df.empty:
                        # Assign columns dynamically based on available columns
                        expected_columns = ['Count', 'Last Name', 'First Name', 'Weight', 'School', 'Grade', 'Rank', 'Gender']
                        df.columns = expected_columns[:len(df.columns)]
                    
                        # Drop rows with missing essential data
                        df = df.dropna(subset=['Last Name'])
                    
                        # Add sheet name to School column
                        df['School'] = sheet
                    
                        # Remove unnecessary columns if they exist
                        df = df.drop(columns=['Grade', 'Count'], errors='ignore')
                    
                        # Normalize Gender column
                        if 'Gender' in df.columns:
                            df['Gender'] = df['Gender'].str.strip().str.lower().map({
                                'boy': 'Boy',
                                'girl': 'Girl'
                            })
                            df = df.dropna(subset=['Gender'])  # Drop rows with invalid Gender values
                    
                        # Add sheet name for reference
                        df['Sheet'] = sheet
                    
                        # Append processed data
                        all_data.append(df)
            
                except Exception as sheet_error:
                    print(f"Error processing sheet '{sheet}': {sheet_error}")
                    continue

            if all_data:
                self.data = pd.concat(all_data, ignore_index=True)
                self.populate_preview()
            else:
                messagebox.showerror("Error", "No valid data found in the file.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")
    # Ayden Winter
    def populate_preview(self):
        # Clear existing data
        for row in self.tree.get_children():
            self.tree.delete(row)

        # Insert new data
        if self.data is not None:
            for _, row in self.data.iterrows():
                self.tree.insert("", tk.END, values=row.tolist())
    #Seth G
    def process_data(self):
        if self.data is None:
            messagebox.showwarning("Warning", "Please select and load a file first.")
        else:
            # Example data processing: Remove duplicates and sort by Weight
            self.data = self.data.drop_duplicates()
            self.data = self.data.sort_values(by=['Gender', 'Weight', 'Rank'])
            self.populate_preview()
            messagebox.showinfo("Info", "Data processed: duplicates removed and sorted by weight.")
    #Seth G
    def edit_selected(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a row to edit.")
            return

        selected_row = self.tree.item(selected_item[0], "values")
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Wrestler")

        labels = ['Last Name', 'First Name', 'Weight', 'School', 'Rank', 'Gender']

        entries = {}

        for i, label in enumerate(labels):
            ttk.Label(edit_window, text=label).grid(row=i, column=0, padx=10, pady=5)
            entry = ttk.Entry(edit_window)
            entry.grid(row=i, column=1, padx=10, pady=5)
            entry.insert(0, selected_row[i])
            entries[label] = entry
        # Ayden Winter
        def save_edit():
            new_values = [entries[label].get() for label in labels]
            # Update the treeview
            self.tree.item(selected_item[0], values=new_values)
            # Update the DataFrame
            self.update_dataframe(selected_row, new_values)
            edit_window.destroy()

        save_btn = ttk.Button(edit_window, text="Save", command=save_edit)
        save_btn.grid(row=len(labels), column=0, columnspan=2, pady=10)
    # Ayden Winter
    def update_dataframe(self, old_values, new_values):
        if self.data is not None:
            # Update the DataFrame row with the new values
            for i, column in enumerate(self.data.columns[:len(new_values)]):
                self.data.loc[(self.data == old_values).all(axis=1), column] = new_values[i]
    # Ayden Winter
    def delete_selected(self):
        selected_items = self.tree.selection()  # Get selected Treeview items
        if not selected_items:
            messagebox.showwarning("Warning", "Please select rows to delete.")
            return

        # Confirm deletion
        if not messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete the selected rows?"):
            return

        # Remove rows from Treeview
        for item in selected_items:
            row_values = self.tree.item(item, "values")  # Get the row values
            self.tree.delete(item)  # Delete the row from the Treeview
        
            # Remove corresponding rows from DataFrame
            if self.data is not None:
                self.data = self.data[~(self.data == list(row_values)).all(axis=1)]

        messagebox.showinfo("Info", "Selected rows have been deleted.")

    # Ayden Winter
    def save_changes(self):
        if self.data is None:
            messagebox.showwarning("Warning", "No data to save.")
        else:
            try:
                save_path = filedialog.asksaveasfilename(
                    title="Save File",
                    defaultextension=".xlsx",
                    filetypes=(
                        ("Excel Files", "*.xlsx"),
                        ("All Files", "*.*")
                    )
                )
                if save_path:
                    self.data.to_excel(save_path, index=False)
                    messagebox.showinfo("Success", f"Data saved to {save_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save data: {e}")

# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = WrestlingBracketApp(root)
    root.mainloop()
